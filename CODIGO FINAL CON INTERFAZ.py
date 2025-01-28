import os
import re
import fitz  # PyMuPDF
import requests
from bs4 import BeautifulSoup
import pandas as pd
import joblib
import PyPDF2
import pdfplumber
import time
from urllib.parse import quote, unquote
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tkinter as tk
from tkinter import messagebox

# Cargar el modelo y el vectorizador
MODELO_PATH = "D:/alejo/MINERIADATOS/CARPETA PROYECTO FINAL/modelo_dispositivos.pkl"
VECTORIZADOR_PATH = "D:/alejo/MINERIADATOS/CARPETA PROYECTO FINAL/vectorizador_dispositivos.pkl"

modelo = joblib.load(MODELO_PATH)
vectorizador = joblib.load(VECTORIZADOR_PATH)

# Configuración
base_url = "https://app.invima.gov.co/alertas/dispositivos-medicos-invima"
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
PDF_FOLDER = "PDFs_Invima"
EXCEL_FILE = "alertas_invima.xlsx"
URL_PATTERNS = [
    re.compile(r"https://app\.invima\.gov\.co/alertas/ckfinder/userfiles/files/ALERTAS%20SANITARIAS/.+\.pdf"),
    re.compile(r"https://app\.invima\.gov\.co/alertas/ckfinder/userfiles/files/INFORMES%20DE%20SEGURIDAD/.+\.pdf")
]
# Funciones combinadas
def limpiar_url(url):
    """Limpia y codifica la URL."""
    url = unquote(url).strip()
    return quote(url, safe=":/")

def obtener_urls(base_url, max_paginas=2):
    """Obtiene todas las URLs de PDFs relevantes desde el sitio web sin repeticiones."""
    
    try:
        urls_encontradas = set()  # Cambiar a un conjunto para evitar duplicados
        pagina_actual = 1
        url_actual = base_url

        while pagina_actual <= max_paginas:
            respuesta = requests.get(url_actual)
            respuesta.raise_for_status()  # Verifica si hubo errores en la solicitud

            # Analiza el contenido HTML de la página actual
            soup = BeautifulSoup(respuesta.text, 'html.parser')

            # Encuentra todos los elementos <a> y extrae sus atributos href
            enlaces = soup.find_all('a', href=True)

            for enlace in enlaces:
                href = enlace['href']
                if href.lower().endswith(".pdf"):
                    full_url = requests.compat.urljoin(base_url, href)
                    # Filtro para incluir solo PDFs de enero de 2025 en adelante
                    if "2025" in full_url:  # Validar si "2025" está en la URL
                        urls_encontradas.add(full_url)  # Añadir al conjunto

                for pattern in URL_PATTERNS:
                    if pattern.match(href):
                        urls_encontradas.add(href)
                        break
                    elif href.startswith("/"):
                        url_absoluta = requests.compat.urljoin(url_actual, href)
                        if pattern.match(url_absoluta):
                            urls_encontradas.add(url_absoluta)
                            break

            siguiente_pagina = soup.find('a', text=re.compile(r"Siguiente", re.IGNORECASE))
            if siguiente_pagina and 'href' in siguiente_pagina.attrs:
                url_actual = requests.compat.urljoin(base_url, siguiente_pagina['href'])
                pagina_actual += 1  # Incrementar número de página
            else:
                break  # Si no hay más páginas, salir del bucle

        # Imprime las URLs encontradas
        for url in urls_encontradas:
            print(url)

        # Devuelve la lista de URLs encontradas (convertida desde el conjunto)
        return list(urls_encontradas)

    except requests.exceptions.RequestException as e:
        return []


def descargar_archivos(urls, carpeta_destino=PDF_FOLDER):
    """Descarga archivos PDF desde una lista de URLs."""
    # Crear carpeta si no existe
    if not os.path.exists(carpeta_destino):
        os.makedirs(carpeta_destino)

    downloaded_files = []  # Lista para almacenar los archivos descargados

    for url in urls:
        try:
            # Limpia y prepara la URL
            url_limpia = limpiar_url(url)
    

            # Nombre del archivo basado en la URL
            filename = os.path.join(carpeta_destino, url_limpia.split('/')[-1])

            respuesta = requests.get(url_limpia, headers=headers, stream=True)
            
            if respuesta.status_code == 500:
                continue
            elif respuesta.status_code == 404:
                continue
            
            respuesta.raise_for_status()
            

            # Guardar el archivo
            with open(filename, 'wb') as archivo:
                for chunk in respuesta.iter_content(chunk_size=8192):
                    archivo.write(chunk)

            downloaded_files.append(filename)  # Agrega el archivo descargado a la lista

            # Espera de 1 segundo entre cada descarga
            time.sleep(1)

        except requests.exceptions.RequestException as e:
            print(f"Error al descargar {url}: {e}")
        except Exception as ex:
            print(f"Error inesperado al procesar {url}: {ex}")

    return downloaded_files

def extract_complete_date(filepath):
    """Extrae la fecha completa del contenido de un PDF."""
    try:
        with fitz.open(filepath) as pdf:
            for page in pdf:
                text = page.get_text()
                # Busca formatos comunes de fecha
                match = re.search(
                    r'\b(\d{1,2})\s(de\s)?(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s(\d{4})\b',
                    text, re.IGNORECASE
                )
                if match:
                    day, _, month, year = match.groups()
                    month_map = {
                        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04", "mayo": "05", "junio": "06",
                        "julio": "07", "agosto": "08", "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12"
                    }
                    formatted_date = f"{int(day):02d}/{month_map[month.lower()]}/{year}"
                    print(f"Fecha completa encontrada: {formatted_date}")
                    return formatted_date
    except Exception as e:
        print(f"Error al procesar el PDF {filepath}: {e}")
    return "Fecha no encontrada"

def extract_alert_info(filepath):
    """Extrae el tipo (Alerta o Informe) y el número asociado del contenido del PDF."""
    try:
        with fitz.open(filepath) as pdf:
            for page in pdf:
                text = page.get_text()
                # Busca los patrones "Alerta No. XXX-XXXX" o "Informe de Seguridad No. XXX-XXXX"
                match = re.search(r'(Alerta|Informe de Seguridad) No\.?\s*(\d{3}-\d{4})', text, re.IGNORECASE)
                if match:
                    alert_type = match.group(1)  # "Alerta" o "Informe de Seguridad"
                    alert_number = match.group(2)  # El número "XXX-XXXX"
                    print(f"Encontrado: {alert_type} - {alert_number}")
                    return alert_type, alert_number
    except Exception as e:
        print(f"Error al procesar el PDF {filepath}: {e}")
    return "Desconocido", "No encontrado"


def extract_month_from_date(date_text):
    """Extrae el mes de una fecha en formato día/mes/año."""
    if date_text == "Fecha no encontrada":
        return "Desconocido"
    try:
        _, month, _ = date_text.split("/")
        month_map = {
            "01": "Enero", "02": "Febrero", "03": "Marzo", "04": "Abril", "05": "Mayo", "06": "Junio",
            "07": "Julio", "08": "Agosto", "09": "Septiembre", "10": "Octubre", "11": "Noviembre", "12": "Diciembre"
        }
        return month_map[month]
    except ValueError:
        return "Desconocido"
    
def extract_device_name(filepath):
    """Extrae el nombre del dispositivo médico o equipo del PDF hasta el siguiente título válido."""
    try:
        with fitz.open(filepath) as pdf:
            for page in pdf:
                text = page.get_text("text")  # Extraer texto plano

                # Verificar si es un Informe de Seguridad
                if "Informe de Seguridad" in text:
                    # Buscar "Asunto" y detener en "No. identificación interna del Informe de Seguridad"
                    match = re.search(
                        r'Asunto\s*:\s*(.+?)(?=\n(?:No\. identificación interna del Informe de Seguridad|[A-ZÁÉÍÓÚÑ ]{2,}[:])|$)',
                        text, re.IGNORECASE | re.DOTALL
                    )
                else:
                    # Buscar "Nombre del producto" para alertas
                    match = re.search(
                        r'Nombre del producto\s*:\s*(.+?)(?=\n(?:No\. identificación interna del Informe de Seguridad|[A-ZÁÉÍÓÚÑ ]{2,}[:])|$)',
                        text, re.IGNORECASE | re.DOTALL
                    )

                if match:
                    device_name = match.group(1).strip()  # Capturar solo el texto relevante
                    return device_name

    except Exception as e:
        print(f"Error al procesar el PDF {filepath}: {e}")

    return "No especificado"

def predecir_tipo_dispositivo(texto):
    """Usa el modelo entrenado para predecir el tipo de dispositivo."""
    try:
        texto_vect = vectorizador.transform([texto])  # Vectorizar el texto
        tipo_predicho = modelo.predict(texto_vect)[0]
        return tipo_predicho
    except Exception as e:
        print(f"Error en la predicción del tipo de dispositivo: {e}")
        return "Predicción fallida"

def extract_registro_invima(filepath):
    """Extrae el número de registro sanitario del contenido del PDF."""
    try:
        with fitz.open(filepath) as pdf:
            for page in pdf:
                text = page.get_text()
                # Busca "Registro sanitario:" seguido por cualquier formato de código relevante
                match = re.search(r'Registro sanitario\s*:\s*(?:INVIMA\s*)?([A-Z0-9-]+)', text, re.IGNORECASE)
                if match:
                    registro = match.group(1)  # Captura el número del registro
                    return registro
    except Exception as e:
        print(f"Error al procesar el PDF {filepath}: {e}")
    return "No encontrado"

def extract_case_description(filepath):
    """
    Extrae el contenido de 'Descripción del caso' de un PDF. 
    Usa PyPDF2 para informes de seguridad y fitz (PyMuPDF) para alertas.
    """
    try:
        # Leer el PDF completo para buscar "Informe de Seguridad"
        with open(filepath, 'rb') as archivo_pdf:
            lector_pdf = PyPDF2.PdfReader(archivo_pdf)
            texto_completo = ""
            for pagina in lector_pdf.pages:
                texto_completo += pagina.extract_text()
        
        # Determinar si es un informe de seguridad o una alerta
        if "Informe de Seguridad" in texto_completo:
            with pdfplumber.open(filepath) as pdf:
                texto_completo = ""
                
                # Leer todas las páginas del PDF
                for pagina in pdf.pages:
                    texto_completo += pagina.extract_text()
            
            # Buscar la sección "Descripción del caso"
            inicio = texto_completo.find("Descripción del caso")
            if inicio != -1:
                # Encontrar el final del apartado
                final = texto_completo.find("Información para profesionales de la salud", inicio)
                if final == -1:
                    final = len(texto_completo)  # Hasta el final del texto si no hay delimitador claro
                
                # Extraer solo el contenido después del título
                contenido = texto_completo[inicio + len("Descripción del caso"):final].strip()
                return contenido
            else:
                return "No se encontró la sección 'Descripción del caso'."  
        else:
            # Lógica para alertas (fitz)
            with fitz.open(filepath) as pdf:
                capturing = False
                description_lines = []

                for page in pdf:
                    text = page.get_text("text")  # Extraer texto de la página
                    lines = text.splitlines()

                    for line in lines:
                        line = line.strip()
                        # Iniciar captura al encontrar "Descripción del caso"
                        if re.match(r'Descripción del caso', line, re.IGNORECASE):
                            capturing = True
                            continue

                        # Detener captura al encontrar un nuevo encabezado
                        if capturing and re.match(r'(Medidas para|Antecedentes|Acciones tomadas|A los|Nota|Referencia|Registro Sanitario|Enlace Relacionado)', line, re.IGNORECASE):
                            capturing = False
                            break

                        # Capturar líneas relevantes
                        if capturing:
                            description_lines.append(line)

                # Unir las líneas capturadas en un solo párrafo
                description = " ".join(description_lines)
                # Limpiar fechas del texto
                description = re.sub(
                    r'\b\d{1,2}\s(de\s)?(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s\d{4}\b',
                    '',
                    description,
                    flags=re.IGNORECASE
                )
                return " ".join(description.split()).strip() if description else "No encontrado"

    except Exception as e:
        print(f"Error al procesar el PDF {filepath}: {e}")
        return "No encontrado"

def update_excel(downloaded_files, excel_file):
    """Actualiza el archivo Excel con los metadatos de los PDFs, empezando desde la fila 6."""
    data = []
    for file in downloaded_files:
        print(f"Procesando archivo: {file}")
        complete_date = extract_complete_date(file)
        month = extract_month_from_date(complete_date)
        alert_type, alert_number = extract_alert_info(file)  # Desempaquetar la tupla
        device_name = extract_device_name(file)
        tipo_dispositivo = predecir_tipo_dispositivo(device_name)
        registro_invima = extract_registro_invima(file)
        descripcion_caso = extract_case_description(file)
    
        data.append({
            "Mes": month,
            "Fecha Completa": complete_date,
            "Numero de alerta (codigo fuente)": alert_number,  # Solo el número
            "Fuente": "INVIMA",
            "Tipo": alert_type,  # Solo el texto del tipo
            "Dispositivo médico o equipo": device_name,
            "Tipo de dispositivo": tipo_dispositivo,
            "Registro INVIMA": registro_invima,
            "Descripción de la alerta Sanitaria o Informe de Seguridad": descripcion_caso,
            "Responsable de verificación": " ",
            "Medio de socialización": " ",
            "Aplicabilidad": " ",
            "Soporte" : file
        })

    # Convertir los nuevos datos en un DataFrame
    new_df = pd.DataFrame(data)

    # Verificar si el archivo Excel existe
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
        ws = wb.active
        print("Archivo Excel existente cargado.")

        # Leer los datos existentes del archivo Excel
        existing_data = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if any(row):  # Ignorar filas vacías
                existing_data.append(row)

        # Crear un DataFrame con los datos existentes
        if existing_data:
            headers = [cell.value for cell in ws[5]]  # Leer encabezados de la fila 5
            existing_df = pd.DataFrame(existing_data, columns=headers)

            # Eliminar duplicados combinando datos existentes con los nuevos
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            combined_df.drop_duplicates(
                subset=["Numero de alerta (codigo fuente)"], keep="first", inplace=True
            )
        else:
            combined_df = new_df
    else:
        wb = Workbook()
        ws = wb.active
        print("Nuevo archivo Excel creado.")

        # Crear encabezados si el archivo no existía
        headers = list(new_df.columns)
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=5, column=col_num, value=header)  # Encabezados en la fila 5
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.font = Font(bold=True, size=14)

        combined_df = new_df

    fill = PatternFill(start_color="A7C7E7", end_color="A7C7E7", fill_type="solid")

    # Definir el estilo de los bordes (todos los bordes)
    border = Border(
        left=Side(style='thin', color="000000"),  # Borde izquierdo
        right=Side(style='thin', color="000000"),  # Borde derecho
        top=Side(style='thin', color="000000"),  # Borde superior
        bottom=Side(style='thin', color="000000")  # Borde inferior
    )

    # Aplicar estilos a las celdas
    ws.merge_cells("C1:G2")
    ws["C1"] = "HOSPITAL UNIVERSITARIO DEL VALLE \"EVARISTO GARCÍA\" E.S.E"
    ws["C1"].font = Font(bold=True)
    ws["C1"].alignment = Alignment(horizontal="center")
    ws["C1"].fill = fill  # Establecer el fondo azul claro
    ws["C1"].border = border  # Aplicar bordes

    ws.merge_cells("C3:G4")
    ws["C3"] = "PLANTILLA DE GESTIÓN Y REVISIÓN DE ALERTAS SANITARIAS"
    ws["C3"].font = Font(bold=True)
    ws["C3"].alignment = Alignment(horizontal="center")
    ws["C3"].fill = fill  # Establecer el fondo azul claro
    ws["C3"].border = border  # Aplicar bordes

    ws["H1"] = "CÓDIGO:"
    ws["I1"] = "2"
    ws.merge_cells("H3:I4")
    ws["H3"] = "FECHA DE EMISIÓN"
    ws["H2"] = "VERSIÓN:"
    ws["I2"] = "FOR-HUV-HUV-009"
    ws["J2"] = "PÁGINA"
    ws["K2"] = "1"
    ws["L2"] = "DE"
    ws["M2"] = "1"
    ws["J3"] = "DÍA"
    ws["K3"] = "MES"
    ws["L3"] = "AÑO"
    ws["J4"] = "5"
    ws["K4"] = "9"
    ws["L4"] = "2019"

    # Aplicar el fondo azul claro y bordes a las celdas de la sección de códigos, versión, página, etc.
    ws["H1"].fill = fill
    ws["I1"].fill = fill
    ws["H2"].fill = fill
    ws["I2"].fill = fill
    ws["J2"].fill = fill
    ws["K2"].fill = fill
    ws["L2"].fill = fill
    ws["M2"].fill = fill
    ws["J3"].fill = fill
    ws["K3"].fill = fill
    ws["L3"].fill = fill
    ws["J4"].fill = fill
    ws["K4"].fill = fill
    ws["L4"].fill = fill

    # Aplicar bordes a todas las celdas del rango A1:M5
    for row in ws["A1:M5"]:
        for cell in row:
            cell.border = border

    # Limpiar la hoja (opcional, si se quiere reemplazar todo)
    ws.delete_rows(6, ws.max_row)

    # Agregar los datos no duplicados desde la fila 6
    for row in dataframe_to_rows(combined_df, index=False, header=False):
        ws.append(row)

    # Guardar los cambios
    wb.save(excel_file)
    print(f"Datos actualizados en {excel_file}")

# Función para manejar la ejecución principal
def ejecutar_actualizacion():
    try:
        # Obtener todos los enlaces a PDFs directamente
        all_pdf_links = obtener_urls(base_url, max_paginas=2)
        print(f"Enlaces a PDFs encontrados: {len(all_pdf_links)}")
        
        if all_pdf_links:
            # Descargar los PDFs
            downloaded_files = descargar_archivos(all_pdf_links, PDF_FOLDER)
            
            # Actualizar el Excel con la información
            update_excel(downloaded_files, EXCEL_FILE)
            messagebox.showinfo("Éxito", "Actualización completada correctamente.")
        else:
            messagebox.showwarning("Advertencia", "No se encontraron enlaces a PDFs para descargar.")
    except Exception as e:
        messagebox.showerror("Error", f"Se produjo un error: {e}")

# Crear la ventana principal de tkinter
root = tk.Tk()
root.title("Actualización de Datos")
root.geometry("300x150")

# Agregar un botón para ejecutar el proceso
btn_actualizar = tk.Button(root, text="Actualizar Datos", command=ejecutar_actualizacion, bg="blue", fg="white", font=("Arial", 12))
btn_actualizar.pack(pady=50)

# Iniciar el bucle de eventos de tkinter
root.mainloop()